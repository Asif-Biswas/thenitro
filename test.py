import asyncio
import requests
from bs4 import BeautifulSoup
import openpyxl


async def scrape(url):
    response = await asyncio.get_event_loop().run_in_executor(None, requests.get, url)
    soup = BeautifulSoup(response.text, 'html.parser')

    membership_level_container = soup.find(id="FunctionalBlock1_ctl00_ctl00_memberProfile_membershipDetails")
    firstname = soup.find(id="FunctionalBlock1_ctl00_ctl00_memberProfile_MemberForm_memberFormRepeater_ctl00_TextBoxLabel10057865")
    lastname = soup.find(id="FunctionalBlock1_ctl00_ctl00_memberProfile_MemberForm_memberFormRepeater_ctl01_TextBoxLabel10057866")
    city = soup.find(id="FunctionalBlock1_ctl00_ctl00_memberProfile_MemberForm_memberFormRepeater_ctl02_TextBoxLabel10234036")
    state = soup.find(id="FunctionalBlock1_ctl00_ctl00_memberProfile_MemberForm_memberFormRepeater_ctl03_DropDownLabel10234033")
    website = soup.find(id="FunctionalBlock1_ctl00_ctl00_memberProfile_MemberForm_memberFormRepeater_ctl04_TextBoxLabel10234040")
    firm_name = soup.find(id="FunctionalBlock1_ctl00_ctl00_memberProfile_MemberForm_memberFormRepeater_ctl05_TextBoxLabel10247801")
    speciality = soup.find(id="FunctionalBlock1_ctl00_ctl00_memberProfile_MemberForm_memberFormRepeater_ctl07_TextBoxLabel10414425")
    membership = soup.find(id="FunctionalBlock1_ctl00_ctl00_memberProfile_MemberForm_memberFormRepeater_ctl09_BulletedList10232588")

    return {
        "firstname": firstname.text if firstname else "",
        "lastname": lastname.text if lastname else "",
        "city": city.text if city else "",
        "state": state.text if state else "",
        "website": website.text if website else "",
        "firm_name": firm_name.text if firm_name else "",
        "speciality": speciality.text if speciality else "",
        "membership_level": membership_level_container.text if membership_level_container else "",
        "membership": membership.text if membership else ""
    }


async def main():
    attorneys2 = openpyxl.load_workbook("attorneys2.xlsx", read_only=False)
    sheet = attorneys2.active

    all_rows = sheet.max_row
    urls = [sheet.cell(row=i, column=2).value for i in range(2, all_rows + 1)]

    tasks = []
    for url in urls:
        tasks.append(asyncio.ensure_future(scrape(url)))

    results = await asyncio.gather(*tasks)

    for i, data in enumerate(results, start=2):
        sheet.cell(row=i, column=4).value = data["membership_level"]
        sheet.cell(row=i, column=5).value = data["firstname"]
        sheet.cell(row=i, column=6).value = data["lastname"]
        sheet.cell(row=i, column=7).value = data["city"]
        sheet.cell(row=i, column=8).value = data["state"]
        sheet.cell(row=i, column=9).value = data["website"]
        sheet.cell(row=i, column=10).value = data["firm_name"]
        sheet.cell(row=i, column=11).value = data["speciality"]
        sheet.cell(row=i, column=12).value = data["membership"]

    attorneys2.save("attorneys3.xlsx")


if __name__ == "__main__":
    loop = asyncio.get_event_loop()
    loop.run_until_complete(main())
