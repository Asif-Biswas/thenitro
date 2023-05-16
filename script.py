from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import openpyxl

def wait_until_loaded(driver, element_id, text):
    element = driver.find_element("id", element_id)
    while element.text == text:
        time.sleep(5)
        element = driver.find_element("id", element_id)

def get_select_options(element):
    return element.find_elements("tag name", "option")

def process_table_row(row, sheet):
    try:
        tds = row.find_elements("tag name", "td")
        attorney_name = tds[0].find_element("tag name", "a").text
        link = tds[0].find_element("tag name", "a").get_attribute("href")
        law_firm = tds[1].text
        print(attorney_name, link, law_firm)

        last_row = sheet.max_row
        sheet.cell(row=last_row + 1, column=1).value = attorney_name
        sheet.cell(row=last_row + 1, column=2).value = link
        sheet.cell(row=last_row + 1, column=3).value = law_firm
    except:
        print("Error processing row", row.text)

def scrape_data(sheet):
    driver = webdriver.Chrome()
    driver.get("https://www.milliondollaradvocates.com/MEMBER-LIST-REFERRAL-DIRECTORY")

    wait_until_loaded(driver, "membersFound", "Loading...")

    idPagingData = driver.find_element("id", "idPagingData")
    idPagingDataSelectTag = idPagingData.find_element("tag name", "select")
    idPagingDataSelectTagOptions = get_select_options(idPagingDataSelectTag)

    for i in range(len(idPagingDataSelectTagOptions)):
        idPagingData = driver.find_element("id", "idPagingData")
        idPagingDataSelectTag = idPagingData.find_element("tag name", "select")
        idPagingDataSelectTagOptions = get_select_options(idPagingDataSelectTag)
        option = idPagingDataSelectTagOptions[i]
        option.click()
        
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.TAG_NAME, 'option')))
        
        time.sleep(2)
        
        membersTable = driver.find_element("id", "membersTable")
        membersTableTbody = membersTable.find_element("tag name", "tbody")
        membersTableTbodyTrs = membersTableTbody.find_elements("tag name", "tr")

        for tr in membersTableTbodyTrs:
            process_table_row(tr, sheet)

    #time.sleep(35)
    driver.close()

def main():
    wb = openpyxl.load_workbook("attorneys.xlsx", read_only=False)
    sheet = wb.active
    print(sheet)

    scrape_data(sheet)
    wb.save("attorneys2.xlsx")

if __name__ == "__main__":
    main()
